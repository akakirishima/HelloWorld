from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from pathlib import Path

import pytest
from fastapi import HTTPException

from app.db.sqlite_db import SqliteDb
from app.services.note_export_service import create_notes_workbook
from app.store.note_store import NoteStore


def test_notes_crud_and_filters(tmp_path: Path) -> None:
    db = SqliteDb(tmp_path / "local.db")
    try:
        store = NoteStore(root=tmp_path, user_id="admin", sqlite_db=db)
        today = date.today()

        first = store.create_note(
            note_date=today,
            title="論文準備",
            did_today="実験計画を更新",
            future_tasks="明日は比較実験",
        )
        assert first.id == today.isoformat()

        with pytest.raises(HTTPException) as exc_info:
            store.create_note(
                note_date=today,
                title="重複",
                did_today="",
                future_tasks="",
            )
        assert exc_info.value.status_code == 409

        second = store.create_note(
            note_date=today - timedelta(days=1),
            title="会議",
            did_today="日程調整",
            future_tasks="資料共有",
        )
        assert second.id == (today - timedelta(days=1)).isoformat()

        filter_by_text = store.list_notes(q="論文", date_from=None, date_to=None)
        assert len(filter_by_text) == 1
        assert filter_by_text[0].id == today.isoformat()

        filter_by_date = store.list_notes(
            q=None,
            date_from=today,
            date_to=today,
        )
        assert len(filter_by_date) == 1
        assert filter_by_date[0].id == today.isoformat()

        moved = store.update_note(
            note_id=today.isoformat(),
            note_date=today + timedelta(days=1),
            title="論文準備",
            did_today="実験計画を更新",
            future_tasks="明日は比較実験",
        )
        assert moved.id == (today + timedelta(days=1)).isoformat()
        assert store.get_note(today.isoformat()) is None
        assert store.get_note((today + timedelta(days=1)).isoformat()) is not None

        store.delete_note(note_id=(today + timedelta(days=1)).isoformat())
        assert store.get_note((today + timedelta(days=1)).isoformat()) is None

        remaining = store.list_notes(q=None, date_from=None, date_to=None)
        assert len(remaining) == 1
        assert remaining[0].id == (today - timedelta(days=1)).isoformat()
    finally:
        db.close()


def test_notes_are_isolated_per_user(tmp_path: Path) -> None:
    db = SqliteDb(tmp_path / "local.db")
    try:
        admin_store = NoteStore(root=tmp_path, user_id="admin", sqlite_db=db)
        member_store = NoteStore(root=tmp_path, user_id="member", sqlite_db=db)
        today = date.today()

        admin_note = admin_store.create_note(
            note_date=today,
            title="管理者メモ",
            did_today="admin only",
            future_tasks="",
        )
        assert member_store.get_note(admin_note.id) is None

        member_store.create_note(
            note_date=today,
            title="メンバーメモ",
            did_today="member only",
            future_tasks="",
        )
        assert len(admin_store.list_notes(q=None, date_from=None, date_to=None)) == 1
        assert len(member_store.list_notes(q=None, date_from=None, date_to=None)) == 1
    finally:
        db.close()


def test_notes_export_workbook(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    db = SqliteDb(tmp_path / "local.db")
    try:
        store = NoteStore(root=tmp_path, user_id="admin", sqlite_db=db)
        today = date.today()
        store.create_note(
            note_date=today,
            title="論文準備",
            did_today="実験計画を更新",
            future_tasks="明日は比較実験",
        )
        store.create_note(
            note_date=today - timedelta(days=1),
            title="会議",
            did_today="日程調整",
            future_tasks="資料共有",
        )

        workbook_bytes = create_notes_workbook(
            store.list_notes(q=None, date_from=None, date_to=None)
        )
        assert workbook_bytes

        wb = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        ws = wb["日誌一覧"]
        assert ws["A2"].value == today.isoformat()
        assert ws["B2"].value == "論文準備"
        assert ws["C2"].value == "実験計画を更新"
    finally:
        db.close()


def test_notes_reject_old_edits(tmp_path: Path) -> None:
    db = SqliteDb(tmp_path / "local.db")
    try:
        store = NoteStore(root=tmp_path, user_id="admin", sqlite_db=db)
        old_date = date.today() - timedelta(days=20)

        with pytest.raises(HTTPException) as exc_info:
            store.create_note(
                note_date=old_date,
                title="古い日誌",
                did_today="",
                future_tasks="",
            )
        assert exc_info.value.status_code == 403
    finally:
        db.close()

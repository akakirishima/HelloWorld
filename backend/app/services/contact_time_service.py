from __future__ import annotations

import calendar
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import range_boundaries
except ModuleNotFoundError:  # pragma: no cover - optional dependency at runtime
    Workbook = None
    Alignment = Border = Font = Side = None

    def range_boundaries(_: str) -> tuple[int, int, int, int]:
        raise RuntimeError("openpyxl is required to generate contact time workbooks.")

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------

WEEKDAYS_JA = ["月", "火", "水", "木", "金", "土", "日"]

# 2026年度: 4〜12月は2026年、1〜2月は2027年
MONTHS: list[tuple[int, int, str]] = [
    (4,  2026, "4月"),
    (5,  2026, "5月"),
    (6,  2026, "6月"),
    (7,  2026, "7月"),
    (8,  2026, "8月"),
    (9,  2026, "9月"),
    (10, 2026, "10月"),
    (11, 2026, "11月"),
    (12, 2026, "12月"),
    (1,  2027, "1月"),
    (2,  2027, "2月"),
]

TITLE_TEXT = "情報処理システム研究室　(コンタクトタイム報告書)"

# 記入例: 2025年4月のサンプルデータ {date: (開始, 終了, 実働H, 備考)}
EXAMPLE_DATA: dict[date, tuple[str, str, int, str]] = {
    date(2025, 4,  9): ("10:00", "17:00", 4, "研究室配属"),
    date(2025, 4, 10): ("10:00", "17:00", 4, "環境構築"),
    date(2025, 4, 11): ("10:00", "17:00", 4, "先輩の卒論読み"),
    date(2025, 4, 14): ("10:00", "17:00", 4, "先輩の卒論読み"),
    date(2025, 4, 15): ("10:00", "17:00", 4, "英論読み"),
    date(2025, 4, 16): ("10:00", "17:00", 4, "英論読み"),
    date(2025, 4, 17): ("10:00", "17:00", 4, "英論読み"),
    date(2025, 4, 18): ("10:00", "17:00", 4, "英論読み"),
    date(2025, 4, 21): ("10:00", "17:00", 4, "英論発表スライド作成"),
    date(2025, 4, 22): ("10:00", "17:00", 4, "英論発表スライド作成"),
    date(2025, 4, 23): ("10:00", "17:00", 4, "英論発表スライド作成"),
    date(2025, 4, 24): ("10:00", "17:00", 4, "英論発表スライド作成"),
    date(2025, 4, 25): ("10:00", "17:00", 4, "英論発表スライド作成"),
    date(2025, 4, 28): ("10:00", "17:00", 4, "英論発表スライド作成"),
    date(2025, 4, 30): ("10:00", "17:00", 4, "英論発表スライド作成"),
}

# ---------------------------------------------------------------------------
# スタイル
# ---------------------------------------------------------------------------

if Workbook is not None:
    _THIN = Side(style="thin")

    F_TITLE = Font(name="ＭＳ 明朝", size=16, bold=True)
    F_HEADER = Font(name="游ゴシック", size=12, bold=True)
    F_DATA = Font(name="游ゴシック", size=11)
    F_SAT = Font(name="游ゴシック", size=11, color="0070C0")
    F_SUN = Font(name="游ゴシック", size=11, color="FF0000")

    BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=False)
else:
    _THIN = None
    F_TITLE = F_HEADER = F_DATA = F_SAT = F_SUN = None
    BORDER_THIN = ALIGN_C = ALIGN_L = None


def _outer_border(ws, cell_range: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = Border(
                top    = _THIN if r == min_row else Side(style=None),
                bottom = _THIN if r == max_row else Side(style=None),
                left   = _THIN if c == min_col else Side(style=None),
                right  = _THIN if c == max_col else Side(style=None),
            )


def _merge(ws, cell_range: str, value=None, font=None, align=None, border=True) -> None:
    ws.merge_cells(cell_range)
    min_col, min_row, *_ = range_boundaries(cell_range)
    cell = ws.cell(row=min_row, column=min_col)
    if value  is not None: cell.value     = value
    if font   is not None: cell.font      = font
    if align  is not None: cell.alignment = align
    if border:
        _outer_border(ws, cell_range)


def _cell(ws, row: int, col: int, value=None, font=None, align=None, border=True):
    c = ws.cell(row=row, column=col)
    if value is not None: c.value     = value
    if font  is not None: c.font      = font
    if align is not None: c.alignment = align
    if border:            c.border    = BORDER_THIN
    return c


def _write_common_header(ws, student_name: str) -> None:
    ws.row_dimensions[2].height = 36
    _merge(ws, "A2:F2", TITLE_TEXT, F_TITLE, ALIGN_C)

    ws.row_dimensions[4].height = 22
    _cell(ws, 4, 1, "学生氏名", F_HEADER, ALIGN_C)
    _merge(ws, "B4:C4", student_name, F_HEADER, ALIGN_C)


def _write_table_header(ws) -> None:
    for r in (7, 8):
        ws.row_dimensions[r].height = 20

    _merge(ws, "A7:A8", "日付\n(yyyy/mm/dd)", F_HEADER, ALIGN_C)
    _merge(ws, "B7:B8", "曜日",               F_HEADER, ALIGN_C)
    _merge(ws, "C7:D7", "活動時間",           F_HEADER, ALIGN_C)
    _cell(ws, 8, 3, "開始時間", F_HEADER, ALIGN_C)
    _cell(ws, 8, 4, "終了時間", F_HEADER, ALIGN_C)
    _merge(ws, "E7:E8", "実働時間(H)", F_HEADER, ALIGN_C)
    _merge(ws, "F7:F8", "備考",        F_HEADER, ALIGN_C)


def _set_col_widths(ws) -> None:
    widths = {"A": 14, "B": 6, "C": 13, "D": 13, "E": 14, "F": 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_data_rows(
    ws,
    dates: list[date],
    sample: dict[date, tuple[str, str, int, str]] | None = None,
) -> tuple[int, int]:
    start_row = 9
    for i, d in enumerate(dates):
        row = start_row + i
        ws.row_dimensions[row].height = 18

        _cell(ws, row, 1, d.strftime("%Y/%m/%d"), F_DATA, ALIGN_C)

        wd = d.weekday()
        font_wd = F_SAT if wd == 5 else F_SUN if wd == 6 else F_DATA
        _cell(ws, row, 2, WEEKDAYS_JA[wd], font_wd, ALIGN_C)

        entry = sample.get(d) if sample else None
        for col_i, val in enumerate(
            [entry[0], entry[1], entry[2], entry[3]] if entry else [None]*4,
            start=3,
        ):
            _cell(ws, row, col_i, val, F_DATA, ALIGN_C)

    last_data_row = start_row + len(dates) - 1
    return start_row, last_data_row


def _write_footer(ws, start_row: int, last_data_row: int) -> None:
    fr = last_data_row + 2

    _merge(ws, f"B{fr}:B{fr+1}", "合 計",     F_HEADER, ALIGN_C)
    _merge(ws, f"C{fr}:D{fr+1}", "稼働 日(日)", F_HEADER, ALIGN_C)
    _merge(ws, f"E{fr}:E{fr+1}",
           f"=COUNTA(E{start_row}:E{last_data_row})", F_HEADER, ALIGN_C)
    _merge(ws, f"F{fr}:F{fr+1}", "承認",       F_HEADER, ALIGN_C)


def _build_monthly_sheet(
    wb: Workbook,
    year: int,
    month: int,
    sheet_name: str,
    student_name: str,
    sample: dict | None = None,
) -> tuple[int, int]:
    ws = wb.create_sheet(title=sheet_name)
    _set_col_widths(ws)
    _write_common_header(ws, student_name)
    _write_table_header(ws)

    num_days = calendar.monthrange(year, month)[1]
    dates = [date(year, month, d) for d in range(1, num_days + 1)]
    start_row, last_data_row = _write_data_rows(ws, dates, sample)
    _write_footer(ws, start_row, last_data_row)

    return start_row, last_data_row


def _build_summary_sheet(
    wb: Workbook,
    month_info: list[tuple[int, int, str, int, int]],
    student_name: str,
) -> None:
    ws = wb.create_sheet(title="合計日数")
    _set_col_widths(ws)
    _write_common_header(ws, student_name)

    ws.row_dimensions[7].height = 20
    _cell(ws, 7, 3, "月",         F_HEADER, ALIGN_C)
    _cell(ws, 7, 4, "稼働日数(日）", F_HEADER, ALIGN_C)
    _cell(ws, 7, 5, "稼働時間(H)",  F_HEADER, ALIGN_C)

    for i, (month_num, _year, sheet_name, start_row, last_row) in enumerate(month_info):
        row = 8 + i
        ws.row_dimensions[row].height = 18
        _cell(ws, row, 3, month_num, F_DATA, ALIGN_C)
        _cell(ws, row, 4,
              f"=COUNTA('{sheet_name}'!E{start_row}:E{last_row})", F_DATA, ALIGN_C)
        _cell(ws, row, 5,
              f"=SUM('{sheet_name}'!E{start_row}:E{last_row})",    F_DATA, ALIGN_C)

    ws.row_dimensions[19].height = 20
    _cell(ws, 19, 3, "合計",          F_HEADER, ALIGN_C)
    _cell(ws, 19, 4, "=SUM(D8:D18)", F_HEADER, ALIGN_C)
    _cell(ws, 19, 5, "=SUM(E8:E18)", F_HEADER, ALIGN_C)

    _cell(ws, 17, 6, "承認", F_HEADER, ALIGN_C)


def _build_example_sheet(wb: Workbook, student_name: str) -> None:
    ws = wb.create_sheet(title="記入例")
    _set_col_widths(ws)
    _write_common_header(ws, student_name)
    _write_table_header(ws)

    num_days = calendar.monthrange(2025, 4)[1]
    dates = [date(2025, 4, d) for d in range(1, num_days + 1)]
    start_row, last_data_row = _write_data_rows(ws, dates, EXAMPLE_DATA)
    _write_footer(ws, start_row, last_data_row)


# ---------------------------------------------------------------------------
# 公開 API
# ---------------------------------------------------------------------------

def create_contact_time_workbook(student_name: str = "") -> Workbook:
    """
    コンタクトタイム報告書 Workbook を生成して返す。

    Args:
        student_name: 学生氏名（空欄可）

    Returns:
        openpyxl.Workbook
    """
    if Workbook is None:
        raise RuntimeError("openpyxl is required to generate contact time workbooks.")
    wb = Workbook()
    del wb[wb.sheetnames[0]]

    month_info: list[tuple[int, int, str, int, int]] = []
    for month_num, year, sheet_name in MONTHS:
        start_row, last_row = _build_monthly_sheet(
            wb, year, month_num, sheet_name, student_name
        )
        month_info.append((month_num, year, sheet_name, start_row, last_row))

    _build_summary_sheet(wb, month_info, student_name)
    _build_example_sheet(wb, student_name)

    return wb


def save_contact_time_workbook(
    user_id: str,
    student_name: str,
    contact_time_root: Path,
) -> Path:
    """
    コンタクトタイム報告書を生成し NAS の所定ディレクトリに保存する。

    Returns:
        保存先の Path
    """
    if Workbook is None:
        raise RuntimeError("openpyxl is required to generate contact time workbooks.")
    out_dir = contact_time_root / "contact_time" / user_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "情報処理システム研究室_コンタクトタイム_2026年度.xlsx"
    wb = create_contact_time_workbook(student_name=student_name)
    wb.save(out_path)
    return out_path

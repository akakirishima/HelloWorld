"""
月別勤怠表 Excel を生成するサービス。
- タイトル行: MS 明朝 16pt
- それ以外: 游ゴシック
- 実働時間: 計算式なし（セッションの duration_sec から整数時間を出力）
- 備考: その日の日誌タイトル + 今日やったこと
- 月の日数に応じて合計行の行番号を自動調整
"""
from __future__ import annotations

import calendar
import urllib.parse
from datetime import date, datetime, timezone
from io import BytesIO
from zoneinfo import ZoneInfo

from app.models.note import NoteRecord
from app.models.session import SessionRecord

JST = ZoneInfo("Asia/Tokyo")
WEEKDAYS_JP = ["月", "火", "水", "木", "金", "土", "日"]

# 行レイアウト定数
_TITLE_ROW = 1
_META_ROW = 2
_HEADER_ROW = 3
_DAY_START_ROW = 4  # 1日はこの行から始まる


def create_timesheet_workbook(
    year: int,
    month: int,
    display_name: str,
    notes: list[NoteRecord],
    sessions: list[SessionRecord],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ModuleNotFoundError as e:
        raise RuntimeError("openpyxl is required to export workbooks.") from e

    days_in_month = calendar.monthrange(year, month)[1]

    # ── 日付ごとの lookup ────────────────────────────────────
    notes_by_date: dict[str, NoteRecord] = {}
    for note in notes:
        key = note.note_date.isoformat() if hasattr(note.note_date, "isoformat") else str(note.note_date)
        notes_by_date[key] = note

    sessions_by_date: dict[str, list[SessionRecord]] = {}
    for s in sessions:
        jst_date = s.check_in_at.astimezone(JST).date().isoformat()
        sessions_by_date.setdefault(jst_date, []).append(s)

    # ── フォント定義 ──────────────────────────────────────────
    f_title   = Font(name="MS 明朝",  size=16, bold=True)
    f_header  = Font(name="游ゴシック", size=11, bold=True)
    f_body    = Font(name="游ゴシック", size=11)
    f_small   = Font(name="游ゴシック", size=10)
    f_weekend = Font(name="游ゴシック", size=11, color="CC0000")

    a_center  = Alignment(horizontal="center", vertical="center")
    a_left    = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    a_right   = Alignment(horizontal="right",  vertical="center")

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── ワークブック生成 ──────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月"

    # Row 1: タイトル（MS 明朝 16pt）
    ws.merge_cells(start_row=_TITLE_ROW, start_column=1, end_row=_TITLE_ROW, end_column=6)
    cell = ws.cell(row=_TITLE_ROW, column=1, value=f"{year}年{month}月　勤怠表")
    cell.font = f_title
    cell.alignment = a_center
    ws.row_dimensions[_TITLE_ROW].height = 32

    # Row 2: 氏名
    ws.cell(row=_META_ROW, column=1, value="氏名").font = f_header
    name_cell = ws.cell(row=_META_ROW, column=2, value=display_name)
    name_cell.font = f_body

    # Row 3: ヘッダー行
    headers = ["日付", "曜日", "出勤", "退勤", "実働時間", "備考"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=h)
        cell.font = f_header
        cell.alignment = a_center
        cell.border = border_all
    ws.row_dimensions[_HEADER_ROW].height = 18

    # Row 4〜: 日付行
    total_hours = 0
    for day in range(1, days_in_month + 1):
        row = _DAY_START_ROW + day - 1
        cur_date = date(year, month, day)
        date_str  = cur_date.isoformat()
        weekday   = cur_date.weekday()          # 0=月〜6=日
        is_weekend = weekday >= 5

        font_day = f_weekend if is_weekend else f_body

        # A: 日付
        cell_date = ws.cell(row=row, column=1, value=f"{month}/{day}")
        cell_date.font = font_day
        cell_date.alignment = a_center
        cell_date.border = border_all

        # B: 曜日
        cell_wd = ws.cell(row=row, column=2, value=WEEKDAYS_JP[weekday])
        cell_wd.font = font_day
        cell_wd.alignment = a_center
        cell_wd.border = border_all

        # セッションデータ
        day_sessions = sorted(sessions_by_date.get(date_str, []), key=lambda s: s.check_in_at)
        if day_sessions:
            first_in  = day_sessions[0].check_in_at.astimezone(JST)
            last_sess = day_sessions[-1]

            cell_in = ws.cell(row=row, column=3, value=first_in.strftime("%H:%M"))
            cell_in.font = font_day
            cell_in.alignment = a_center
            cell_in.border = border_all

            if last_sess.check_out_at is not None:
                last_out = last_sess.check_out_at.astimezone(JST)
                cell_out = ws.cell(row=row, column=4, value=last_out.strftime("%H:%M"))
            else:
                cell_out = ws.cell(row=row, column=4, value="未退勤")
            cell_out.font = font_day
            cell_out.alignment = a_center
            cell_out.border = border_all

            total_sec = sum(s.duration_sec for s in day_sessions if s.duration_sec is not None)
            hours = int(total_sec // 3600)
            cell_h = ws.cell(row=row, column=5, value=hours)
            cell_h.font = font_day
            cell_h.alignment = a_center
            cell_h.border = border_all
            total_hours += hours
        else:
            for col in (3, 4, 5):
                ws.cell(row=row, column=col).border = border_all

        # F: 備考（日誌）
        note = notes_by_date.get(date_str)
        parts: list[str] = []
        if note:
            if note.title:
                parts.append(note.title)
            if note.did_today:
                parts.append(note.did_today)
        cell_note = ws.cell(row=row, column=6, value="\n".join(parts) if parts else None)
        cell_note.font = f_small
        cell_note.alignment = a_left
        cell_note.border = border_all

        # 備考あり行は高さを広げる
        if parts:
            lines = sum(v.count("\n") + 1 for v in parts)
            ws.row_dimensions[row].height = max(15, min(lines * 15, 90))

    # 合計行（位置 = _DAY_START_ROW + days_in_month）
    total_row = _DAY_START_ROW + days_in_month
    ws.cell(row=total_row, column=1, value="合計").font   = f_header
    ws.cell(row=total_row, column=1).alignment            = a_center
    ws.cell(row=total_row, column=1).border               = border_all
    for col in (2, 3, 4):
        ws.cell(row=total_row, column=col).border = border_all
    cell_total = ws.cell(row=total_row, column=5, value=total_hours)
    cell_total.font      = f_header
    cell_total.alignment = a_center
    cell_total.border    = border_all
    ws.cell(row=total_row, column=6).border = border_all

    # 承認行
    approval_row = total_row + 1
    ws.cell(row=approval_row, column=1, value="承認").font      = f_header
    ws.cell(row=approval_row, column=1).alignment               = a_center
    ws.cell(row=approval_row, column=1).border                  = border_all
    for col in range(2, 7):
        ws.cell(row=approval_row, column=col).border = border_all
    ws.row_dimensions[approval_row].height = 24

    # 列幅
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width =  6
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 46

    ws.freeze_panes = "A4"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def timesheet_filename(year: int, month: int) -> str:
    name = f"勤怠表_{year}年{month:02d}月.xlsx"
    return urllib.parse.quote(name, safe="")
